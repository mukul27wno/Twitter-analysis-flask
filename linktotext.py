import pandas as pd
from openpyxl import load_workbook
from textblob import TextBlob
import re
import nltk
from nltk.corpus import stopwords
import requests
from bs4 import BeautifulSoup
import pathlib
import os

inputurl = pd.read_excel('Input.xlsx')
headers = {
    "User-Agent": "Mozilla/5.0 (X11; Linux x86_64; rv:60.0) Gecko/20100101 Firefox/60.0"
}
url = inputurl['URL']


cloc = pathlib.Path().resolve()
directory = 'txtfiles'
path = os.path.join(cloc, directory)
os.mkdir(path)

for i in range(len(url)):
  raw_html = requests.get(url[i], headers=headers)
  soup = BeautifulSoup(raw_html.content, 'html.parser')
  data = ''
  fp = open(f'txtfiles/{i+1}.txt', 'w+')
  for data in soup.find_all("p"):
      fp.write(data.get_text())
  fp.close()


def stopwordss(wordscopy):
    for x in wordscopy:
        return [x for x in wordscopy if x not in stopwords.words('english')]


def analysiss(text, i):
    pl = TextBlob(text).sentiment.polarity
    su = TextBlob(text).sentiment.subjectivity
    if pl <= 0:
        pos = -1 - pl
        neg = 1 + pl
    else:
        pos = 1 - pl
        neg = -1 + pl

    sentences = text.split(".")
    sumsen = 0
    tns = 0
    for y in sentences:
        tns = tns + 1  # Total Number of Sentences

    words = text.split(" ")
    tnw = 0
    wordscopy = words

    wordscopy = stopwordss(wordscopy)

    punctuations = '''!()-[]{};:'"\,<>./?@#$%^&*_~'''

    wordscopy2 = ""
    for char in wordscopy:
        if char not in punctuations:
            wordscopy2 = wordscopy2 + char

    for x in wordscopy:
        tnw = tnw + 1  # Total number of words

    ccw = 0
    cwp = ['un', 'non', 'in', 'pre', 'trans', 're', 'con']
    cws = ['ly', 'ist', 'er', 'ness', 'ment', 's', 'ing', 'ed', 'en', 'est', 'mit', 'ceive', 'fer']
    for y in words:
        for y2 in cwp:
            if y.startswith(y2):
                ccw = ccw + 1
        for y2 in cws:
            if y.endswith(y2):
                ccw = ccw + 1

    pcw = ccw / tnw

    fi = 0.4 * ((tnw / tns) + 100 * (ccw / tnw))

    anws = tnw / tns

    sc = 0
    for w in wordscopy2:
        if (
                w == 'a' or w == 'e' or w == 'i' or w == 'o' or w == 'u' or w == 'A' or w == 'E' or w == 'I' or w == 'O' or w == 'U'):
            sc = sc + 1

    scw = sc / tnw

    pronounRegex = re.compile(
        r'\b(I|me|mine|myself|us|our|ourselves|you|your|yours|yourself|yourselves|he|him|himself|his|she|her|hers|herself|it|its|itself|they|them|their|theirs|themselves|we|my|ours|(?-i:us))\b',
        re.I)
    pronouns = pronounRegex.findall(text)

    asl = len(wordscopy2) / tns
    awl = len(wordscopy2) / tnw

    poscell = my_sheet_obj.cell(row=i, column=3)
    poscell.value = pos
    negcell = my_sheet_obj.cell(row=i, column=4)
    negcell.value = neg
    pscell = my_sheet_obj.cell(row=i, column=5)
    pscell.value = pl
    sucell = my_sheet_obj.cell(row=i, column=6)
    sucell.value = su
    aslcell = my_sheet_obj.cell(row=i, column=7)
    aslcell.value = asl
    pcwcell = my_sheet_obj.cell(row=i, column=8)
    pcwcell.value = pcw
    ficell = my_sheet_obj.cell(row=i, column=9)
    ficell.value = fi
    anwscell = my_sheet_obj.cell(row=i, column=10)
    anwscell.value = anws
    cnccell = my_sheet_obj.cell(row=i, column=11)
    cnccell.value = ccw
    wccell = my_sheet_obj.cell(row=i, column=12)
    wccell.value = tnw
    sccell = my_sheet_obj.cell(row=i, column=13)
    sccell.value = scw
    ppcell = my_sheet_obj.cell(row=i, column=14)
    ppcell.value = len(pronouns)
    awlcell = my_sheet_obj.cell(row=i, column=15)
    awlcell.value = awl

my_wb_obj = load_workbook(filename="Output Data Structure.xlsx")
my_sheet_obj = my_wb_obj.active
for i in range(len(url)):
  tf = open(f'txtfiles/{i+1}.txt','r')
  analysiss(tf.read(),i+2)
  tf.close()
my_wb_obj.save("Output Data Structure.xlsx")import pandas as pd
from openpyxl import load_workbook
from textblob import TextBlob
import re
import nltk
from nltk.corpus import stopwords
import requests
from bs4 import BeautifulSoup
import pathlib
import os

inputurl = pd.read_excel('Input.xlsx')
headers = {
    "User-Agent": "Mozilla/5.0 (X11; Linux x86_64; rv:60.0) Gecko/20100101 Firefox/60.0"
}
url = inputurl['URL']


cloc = pathlib.Path().resolve()
directory = 'txtfiles'
path = os.path.join(cloc, directory)
os.mkdir(path)

for i in range(len(url)):
  raw_html = requests.get(url[i], headers=headers)
  soup = BeautifulSoup(raw_html.content, 'html.parser')
  data = ''
  fp = open(f'txtfiles/{i+1}.txt', 'w+')
  for data in soup.find_all("p"):
      fp.write(data.get_text())
  fp.close()


def stopwordss(wordscopy):
    for x in wordscopy:
        return [x for x in wordscopy if x not in stopwords.words('english')]


def analysiss(text, i):
    pl = TextBlob(text).sentiment.polarity
    su = TextBlob(text).sentiment.subjectivity
    if pl <= 0:
        pos = -1 - pl
        neg = 1 + pl
    else:
        pos = 1 - pl
        neg = -1 + pl

    sentences = text.split(".")
    sumsen = 0
    tns = 0
    for y in sentences:
        tns = tns + 1  # Total Number of Sentences

    words = text.split(" ")
    tnw = 0
    wordscopy = words

    wordscopy = stopwordss(wordscopy)

    punctuations = '''!()-[]{};:'"\,<>./?@#$%^&*_~'''

    wordscopy2 = ""
    for char in wordscopy:
        if char not in punctuations:
            wordscopy2 = wordscopy2 + char

    for x in wordscopy:
        tnw = tnw + 1  # Total number of words

    ccw = 0
    cwp = ['un', 'non', 'in', 'pre', 'trans', 're', 'con']
    cws = ['ly', 'ist', 'er', 'ness', 'ment', 's', 'ing', 'ed', 'en', 'est', 'mit', 'ceive', 'fer']
    for y in words:
        for y2 in cwp:
            if y.startswith(y2):
                ccw = ccw + 1
        for y2 in cws:
            if y.endswith(y2):
                ccw = ccw + 1

    pcw = ccw / tnw

    fi = 0.4 * ((tnw / tns) + 100 * (ccw / tnw))

    anws = tnw / tns

    sc = 0
    for w in wordscopy2:
        if (
                w == 'a' or w == 'e' or w == 'i' or w == 'o' or w == 'u' or w == 'A' or w == 'E' or w == 'I' or w == 'O' or w == 'U'):
            sc = sc + 1

    scw = sc / tnw

    pronounRegex = re.compile(
        r'\b(I|me|mine|myself|us|our|ourselves|you|your|yours|yourself|yourselves|he|him|himself|his|she|her|hers|herself|it|its|itself|they|them|their|theirs|themselves|we|my|ours|(?-i:us))\b',
        re.I)
    pronouns = pronounRegex.findall(text)

    asl = len(wordscopy2) / tns
    awl = len(wordscopy2) / tnw

    poscell = my_sheet_obj.cell(row=i, column=3)
    poscell.value = pos
    negcell = my_sheet_obj.cell(row=i, column=4)
    negcell.value = neg
    pscell = my_sheet_obj.cell(row=i, column=5)
    pscell.value = pl
    sucell = my_sheet_obj.cell(row=i, column=6)
    sucell.value = su
    aslcell = my_sheet_obj.cell(row=i, column=7)
    aslcell.value = asl
    pcwcell = my_sheet_obj.cell(row=i, column=8)
    pcwcell.value = pcw
    ficell = my_sheet_obj.cell(row=i, column=9)
    ficell.value = fi
    anwscell = my_sheet_obj.cell(row=i, column=10)
    anwscell.value = anws
    cnccell = my_sheet_obj.cell(row=i, column=11)
    cnccell.value = ccw
    wccell = my_sheet_obj.cell(row=i, column=12)
    wccell.value = tnw
    sccell = my_sheet_obj.cell(row=i, column=13)
    sccell.value = scw
    ppcell = my_sheet_obj.cell(row=i, column=14)
    ppcell.value = len(pronouns)
    awlcell = my_sheet_obj.cell(row=i, column=15)
    awlcell.value = awl

my_wb_obj = load_workbook(filename="Output Data Structure.xlsx")
my_sheet_obj = my_wb_obj.active
for i in range(len(url)):
  tf = open(f'txtfiles/{i+1}.txt','r')
  analysiss(tf.read(),i+2)
  tf.close()
my_wb_obj.save("Output Data Structure.xlsx")